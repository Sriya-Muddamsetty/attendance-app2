import os
import cv2
import numpy as np
from flask import Flask, render_template, jsonify, request
from tensorflow.keras.models import load_model
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# Initialize Flask app
app = Flask(__name__)

# Load the trained model
model = load_model("C:/Users/megha/OneDrive/Desktop/attendance-app/class_roll_number_model.h5")

# Load the class labels (roll numbers) from the saved file
class_names = np.load('classes.npy', allow_pickle=True)

# Sort the class names
sorted_class_names = sorted(class_names)

# Define the image size expected by the model
img_size = (100, 100)

# Excel file setup
excel_file = 'Student_Attendance.xlsx'
def setup_excel():
    if not os.path.exists(excel_file):
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Title at the top
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = "Student Attendance System"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Add column headers
        ws.cell(row=2, column=1, value="Roll Number")
        ws.cell(row=2, column=2, value="Date")

        # Add class names to the first column
        for idx, class_name in enumerate(sorted_class_names, start=3):
            ws.cell(row=idx, column=1, value=class_name)

        # Save the workbook
        wb.save(excel_file)

setup_excel()

@app.route('/')
def index():
    # Render the unified HTML file
    return render_template('index.html')

@app.route('/view_attendance_data', methods=['GET'])
def view_attendance_data():
    # Load attendance data from the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active
    data = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        data.append(row)

    # Send data as JSON to populate the View Attendance section
    return jsonify(data)

@app.route('/predict', methods=['POST'])
def predict():
    selected_date = request.form.get('date')

    # Check if the image is sent
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400

    # Get the image from the request
    img_file = request.files['image']

    # Convert the image to numpy array
    img_array = np.array(bytearray(img_file.read()), dtype=np.uint8)

    # Convert byte array to OpenCV image
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    if img is None:
        return jsonify({'error': 'Failed to decode image'}), 400

    # Resize the image to the model's expected input size
    img = cv2.resize(img, img_size)

    # Normalize the image
    img = img / 255.0

    # Prepare image for prediction (add batch dimension)
    img_array = np.expand_dims(img, axis=0)

    # Make prediction
    prediction = model.predict(img_array)

    # Get the predicted label
    predicted_label = np.argmax(prediction, axis=1)
    predicted_label = int(predicted_label[0])

    # Map the predicted label back to the roll number (class name)
    roll_number = class_names[predicted_label]

    # Update the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active

    # Check if the roll number is already marked present for the date
    col_idx = None
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=2, column=col).value == selected_date:
            col_idx = col
            break

    if col_idx is None:
        col_idx = ws.max_column + 1
        ws.cell(row=2, column=col_idx, value=selected_date)

    already_present = False
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == roll_number:
            if ws.cell(row=row, column=col_idx).value == "Present":
                already_present = True
            else:
                ws.cell(row=row, column=col_idx, value="Present")
            break

    # Count the number of "Present" entries for the date
    present_count = sum(1 for row in range(3, ws.max_row + 1) if ws.cell(row=row, column=col_idx).value == "Present")

    # Write the count at the end of the sheet
    ws.cell(row=ws.max_row + 1, column=col_idx, value=f"Total Present: {present_count}")

    # Save the workbook
    wb.save(excel_file)

    if already_present:
        return jsonify({'roll_number': roll_number, 'message': 'Already taken'}), 200

    return jsonify({'roll_number': roll_number}), 200

if __name__ == '__main__':
    # Point Flask to the templates folder
    app.template_folder = 'C:\\Users\\megha\\OneDrive\\Desktop\\attendance-app\\templates'
    app.run(debug=True)
