import os
import cv2
import numpy as np
from flask import Flask, render_template, jsonify, request
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing import image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# Initialize Flask app
app = Flask(__name__)

# Load the trained model
model = load_model('class_roll_number_model.h5')

# Load the class labels (roll numbers) from the saved file
class_names = np.load('classes.npy', allow_pickle=True)

# Sort the class names
sorted_class_names = sorted(class_names)

# Define the image size expected by the model
img_size = (100, 100)

# Excel file setup
excel_file = 'Student_Attendance.xlsx'
def setup_excel():
    if not os.path.exists(excel_file):
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Title at the top
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = "Student Attendance System"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Add class names to the first column
        for idx, class_name in enumerate(sorted_class_names, start=2):
            ws.cell(row=idx, column=1, value=class_name)

        # Save the workbook
        wb.save(excel_file)

setup_excel()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/predict', methods=['POST'])
def predict():
    # Check if the image is sent
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400

    # Get the image from the request
    img_file = request.files['image']

    # Convert the image to numpy array
    img_array = np.array(bytearray(img_file.read()), dtype=np.uint8)

    # Convert byte array to OpenCV image
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    if img is None:
        return jsonify({'error': 'Failed to decode image'}), 400

    # Resize the image to the model's expected input size
    img = cv2.resize(img, img_size)

    # Normalize the image
    img = img / 255.0

    # Prepare image for prediction (add batch dimension)
    img_array = np.expand_dims(img, axis=0)

    # Make prediction
    prediction = model.predict(img_array)

    # Get the predicted label
    predicted_label = np.argmax(prediction, axis=1)
    predicted_label = int(predicted_label[0])

    # Map the predicted label back to the roll number (class name)
    roll_number = class_names[predicted_label]

    # Update the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active

    # Find the row corresponding to the roll number and mark "Present"
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == roll_number:
            ws.cell(row=row, column=2, value="Present")

    # Count the number of "Present" entries
    present_count = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=2).value == "Present")

    # Write the count at the end of the sheet
    ws.cell(row=ws.max_row + 1, column=1, value="Total Present")
    ws.cell(row=ws.max_row, column=2, value=present_count)

    # Save the workbook
    wb.save(excel_file)

    # Return the predicted roll number as JSON
    return jsonify({'roll_number': roll_number})

if __name__ == '__main__':
    app.run(debug=True)
